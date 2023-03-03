import re
from django.http import HttpResponse
from rest_framework import status
from rest_framework.decorators import api_view
from rest_framework.response import Response
import openpyxl


@api_view(['GET'])
def check_duplicate(request, sentence):
    for i in sentence.split():
        for ii in i:
            if i.count(ii) > 1:
                return Response(True, status=status.HTTP_200_OK)
    return Response(False, status=status.HTTP_200_OK)


@api_view(['POST'])
def commas(request):
    if type(request.data) != int:
        return Response("Please pass correct number", status=status.HTTP_400_BAD_REQUEST)
    return Response(f'{request.data:,}', status=status.HTTP_200_OK)


@api_view(['POST'])
def upload_file(request):
    """
    List all code snippets, or create a new snippet.
    """
    file_obj = request.data['file']
    xfile = openpyxl.load_workbook(file_obj, keep_vba=True, read_only=True)
    sheet = xfile.worksheets[0]
    max_row = sheet.max_row
    lst = []
    for row in range(2, max_row + 1):
        if sheet.cell(row=row, column=1).value == None:
            break
        print(re.findall(r'"(.*?)"', str(sheet.cell(row=row,
              column=1).value).split()[0])[0], "QQPP!!")
        print(sheet.cell(row=row, column=1).value)
        lst.append(
            {
                "title": re.findall(r'"(.*?)"', str(sheet.cell(row=row, column=1).value).split()[0])[0],
                "description": re.findall(r'"(.*?)"', str(sheet.cell(row=row, column=1).value).split("=")[2].split(',')[0])[0],
                "tone": re.findall(r'"(.*?)"', str(sheet.cell(row=row, column=1).value).split()[-1])[0],
                "numModel1": len(str(sheet.cell(row=row, column=2).value)),
                "numModel2": len(str(sheet.cell(row=row, column=3).value))
            }


        )
    return Response({"data": lst}, status=status.HTTP_200_OK)
